from openpyxl import load_workbook, Workbook


def xlsx_read(file, max_col):
    sheet_data = []
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_col=1, max_col=max_col, values_only=True):
        sheet_data.append(tuple(row))
    return sheet_data


def xlsx_write(file_name, data):
    wb = Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(file_name)
